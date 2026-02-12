"""Xindus UAT API client — authenticate and submit B2B shipments.

Uses the Express-Shipment API (multipart: Excel + JSON) which creates
proper B2B shipments (b2b_shipment=1, source=B2B).
"""
from __future__ import annotations

import json
import logging
import re
import time
from io import BytesIO
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.config import XINDUS_UAT_URL, XINDUS_UAT_USERNAME, XINDUS_UAT_PASSWORD

logger = logging.getLogger(__name__)

# In-memory token cache
_token: str | None = None
_token_expires: float = 0


async def _authenticate() -> str:
    """Login to Xindus UAT and return a Bearer token (cached)."""
    global _token, _token_expires

    if _token and time.time() < _token_expires:
        return _token

    if not XINDUS_UAT_USERNAME or not XINDUS_UAT_PASSWORD:
        raise RuntimeError("XINDUS_UAT_USERNAME / XINDUS_UAT_PASSWORD not configured")

    url = f"{XINDUS_UAT_URL}/xos/api/auth/login"
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(url, json={
            "username": XINDUS_UAT_USERNAME,
            "password": XINDUS_UAT_PASSWORD,
        })

    if resp.status_code != 200:
        raise RuntimeError(f"Xindus auth failed ({resp.status_code}): {resp.text[:200]}")

    data = resp.json()
    # Token is nested: {"data": [{"access_token": "..."}]}
    token_data = (data.get("data") or [{}])[0] if isinstance(data.get("data"), list) else data
    _token = token_data.get("access_token") or data.get("access_token") or data.get("token")
    if not _token:
        raise RuntimeError(f"No access_token in auth response: {data}")

    # Cache for 55 minutes (tokens typically last 1h)
    _token_expires = time.time() + 55 * 60
    logger.info("Xindus UAT auth successful, token cached")
    return _token


def _clear_token() -> None:
    """Clear cached token (used on 401 retry)."""
    global _token, _token_expires
    _token = None
    _token_expires = 0


def _normalize_hs(val: Any) -> str | None:
    """Normalize an HS/HTS code string for Excel.

    Keeps as STRING to preserve leading zeros — Xindus Java parser
    uses getCellString() which handles both STRING and NUMERIC cells.
    Export HSN (ehsn) must be exactly 8 characters for validation.
    Returns None for empty/missing values (cell left blank).
    """
    if val is None or val == "":
        return None
    s = str(val).replace(".", "").replace(" ", "").strip()
    return s if s else None


def _camel_to_snake(name: str) -> str:
    """Convert a camelCase key to snake_case.

    Examples: shippingMethod → shipping_method,
              amazonFba → amazon_fba,
              extensionNumber → extension_number.
    Single-word keys (name, email, country) pass through unchanged.
    """
    s = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1_\2", name)
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s)
    return s.lower()


def _to_snake_keys(obj: Any) -> Any:
    """Recursively convert all dict keys from camelCase to snake_case.

    The Xindus Express-Shipment DTO uses @JsonProperty with snake_case
    names (e.g. "origin_clearance_type", "shipper_address").  Our frontend
    and draft storage use camelCase.  This function bridges the gap.
    """
    if isinstance(obj, dict):
        return {_camel_to_snake(k): _to_snake_keys(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_to_snake_keys(item) for item in obj]
    return obj


def _build_excel(shipment_data: dict[str, Any]) -> bytes:
    """Generate XpressB2B Excel for Xindus Express-Shipment API.

    Headers, sheet name, styling, and layout match the official Xindus
    "FBA Split shipment" template.
    - Single-address (multiAddressDestinationDelivery=false): 12 columns
    - Multi-address  (multiAddressDestinationDelivery=true):  21 columns
    First item row of each box has all box-level columns filled;
    subsequent items leave box-level columns empty (None).
    """
    # Detect mode from the payload (camelCase key from frontend)
    multi_addr = shipment_data.get("multiAddressDestinationDelivery", False)
    if not multi_addr:
        multi_addr = shipment_data.get("multi_address_destination_delivery", False)

    # ── Styles matching the official Xindus template ──
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hdr_fill_req = PatternFill("solid", fgColor="073763")   # dark blue (required)
    hdr_fill_opt = PatternFill("solid", fgColor="6FA8DC")   # light blue (optional)
    hdr_align = Alignment(horizontal="left", vertical="center")
    hdr_align_center = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center")

    # ── Column definitions: (header, width, optional?) ──
    if multi_addr:
        col_defs: list[tuple[str, float, bool]] = [
            ("Box Number (R)",              15.0, False),
            ("Receiver Name (R)",           17.5, False),
            ("Receiver Address (R)",        28.0, False),
            ("Receiver City (R)",           15.0, False),
            ("Receiver Zip (R)",            12.0, False),
            ("Receiver State (R)",          15.0, False),
            ("Receiver Country (R)",        16.0, False),
            ("Receiver Phone Number (R)",   22.0, False),
            ("Receiver Extension No (O)",   22.0, True),
            ("Receiver Email (R)",          20.0, False),
            ("Box Length(cms)(R)",           18.0, False),
            ("Box Width(cms)(R)",           17.0, False),
            ("Box Height(cms)(R)",          18.0, False),
            ("Box Weight(kgs)(R)",          18.0, False),
            ("Item Description (R)",        25.0, False),
            ("Item Qty (R)",                12.0, False),
            ("Item Unit Weight Kg (O)",     22.0, True),
            ("HS Code(Origin)(R)",          18.0, False),
            ("HS Code(Destination) (O)",    22.0, True),
            ("Item Unit Price (R)",         18.0, False),
            ("IGST % (For GST taxtype)",    22.0, True),
        ]
        box_col_count = 14  # columns 1-14 are box-level
    else:
        col_defs = [
            ("Box Number (R)",              20.0, False),
            ("Box Length(cms)(R)",           19.0, False),
            ("Box Width(cms)(R)",           18.0, False),
            ("Box Height(cms)(R)",          19.0, False),
            ("Box Weight(kgs)(R)",          19.0, False),
            ("Item Description (R)",        25.0, False),
            ("Item Qty (R)",                12.0, False),
            ("Item Unit Weight Kg (O)",     22.0, True),
            ("HS Code(Origin)(R)",          18.0, False),
            ("HS Code(Destination) (O)",    22.0, True),
            ("Item Unit Price (R)",         18.0, False),
            ("IGST % (For GST taxtype)",    22.0, True),
        ]
        box_col_count = 5  # columns 1-5 are box-level

    columns = [d[0] for d in col_defs]

    wb = Workbook()
    ws = wb.active
    ws.title = "FBA Split shipment"

    # ── Header row with Xindus styling ──
    ws.row_dimensions[1].height = 16.5
    for ci, (hdr, width, is_opt) in enumerate(col_defs, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill_opt if is_opt else hdr_fill_req
        cell.alignment = hdr_align_center if ci == 1 else hdr_align
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Fallback receiver from top-level
    top_recv = shipment_data.get("receiverAddress") or shipment_data.get("receiver_address") or {}
    boxes = (shipment_data.get("shipmentBoxes")
             or shipment_data.get("shipment_boxes")
             or [])

    current_row = 2
    for box in boxes:
        items = (box.get("shipmentBoxItems")
                 or box.get("shipment_box_items")
                 or [{}])

        box_id = box.get("boxId", box.get("box_id", 0))

        for j, item in enumerate(items):
            is_first = j == 0

            if multi_addr:
                recv = box.get("receiverAddress") or box.get("receiver_address") or {}
                if not recv.get("name"):
                    recv = top_recv
                if is_first:
                    row_data: list[Any] = [
                        box_id,
                        recv.get("name", ""),
                        recv.get("address", ""),
                        recv.get("city", ""),
                        recv.get("zip", ""),
                        recv.get("state", ""),
                        recv.get("country", ""),
                        recv.get("phone", ""),
                        recv.get("extensionNumber", recv.get("extension_number", "")) or None,
                        recv.get("email", ""),
                        box.get("length", 0),
                        box.get("width", 0),
                        box.get("height", 0),
                        box.get("weight", 0),
                    ]
                else:
                    row_data = [None] * box_col_count
            else:
                if is_first:
                    row_data = [
                        box_id,
                        box.get("length", 0),
                        box.get("width", 0),
                        box.get("height", 0),
                        box.get("weight", 0),
                    ]
                else:
                    row_data = [None] * box_col_count

            # Item-level columns (same for both formats).
            # HS codes kept as STRINGS to preserve leading zeros —
            # Xindus getCellString() handles both STRING and NUMERIC cells,
            # but ehsn validation requires exactly 8 characters.
            row_data.extend([
                item.get("description", "") or None,
                item.get("quantity", 0),
                item.get("weight", 0) or None,
                _normalize_hs(item.get("ehsn", "")),
                _normalize_hs(item.get("ihsn", "")),
                item.get("unitPrice", item.get("unit_price", 0)),
                item.get("igst", item.get("igst_amount", 0)) or None,
            ])

            for ci, val in enumerate(row_data, start=1):
                if val is None:
                    continue
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.font = data_font
                cell.alignment = data_align

            current_row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def upload_document(
    file_data: bytes,
    filename: str,
    docs_type: str = "FULFIL_LABEL",
) -> str:
    """Upload a file to Xindus CDN via POST /xos/api/file.

    Returns the CDN URL string. Xindus stores files at ucdn.xindus.net.
    """
    token = await _authenticate()
    url = f"{XINDUS_UAT_URL}/xos/api/file"

    files = [("file", (filename, file_data, "application/pdf"))]
    data_fields = {
        "docs_type": docs_type,
        "reference_string": "0",
        "is_private": "false",
    }

    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            url,
            files=files,
            data=data_fields,
            headers={"Authorization": f"Bearer {token}"},
        )

    if resp.status_code == 401:
        # Retry with fresh token
        _clear_token()
        token = await _authenticate()
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(
                url,
                files=files,
                data=data_fields,
                headers={"Authorization": f"Bearer {token}"},
            )

    if resp.status_code not in (200, 201):
        raise RuntimeError(f"Xindus file upload failed ({resp.status_code}): {resp.text[:300]}")

    body = resp.json()
    raw_data = body.get("data")
    if isinstance(raw_data, list) and raw_data:
        cdn_url = raw_data[0]
        logger.info("Xindus file upload OK: %s → %s", filename, cdn_url)
        return cdn_url
    raise RuntimeError(f"No URL in Xindus file upload response: {body}")


async def submit_b2b_shipment(
    shipment_data: dict[str, Any],
    consignor_id: int | None = None,
) -> tuple[int, dict[str, Any]]:
    """Submit a B2B shipment to Xindus UAT via the Express-Shipment API.

    Xindus B2B creation is a two-step process:
      1. **Initiate** — POST /initiate with ``initiate_shipment_data``
         Creates a shipment entry and returns a ``scancode``.
      2. **Create** — POST /create with ``create_shipment_data``
         Updates the initiated shipment (looked up by scancode) with
         full details, charges, docs, etc.

    Returns (http_status, response_body).
    Retries once on 401 with a fresh token.
    """
    # 1. Generate Excel from shipment boxes (handles both camelCase and snake_case)
    excel_bytes = _build_excel(shipment_data)

    # 2. Convert camelCase keys → snake_case for Xindus DTO.
    #    The Java B2BShipmentCreateRequestDTO uses @JsonProperty with snake_case
    #    names (e.g. "origin_clearance_type"), so camelCase keys are silently
    #    ignored, leaving fields null and causing NPE in parseShipmentItem().
    snake_data = _to_snake_keys(shipment_data)

    base_url = f"{XINDUS_UAT_URL}/xos/api/express-shipment"

    for attempt in range(2):
        token = await _authenticate()
        headers = {"Authorization": f"Bearer {token}"}

        # ── Step 1: Initiate — create shipment entry, get scancode ──
        initiate_payload = json.dumps(snake_data)
        initiate_files = [
            ("box_details_file", ("uploadedFile.xlsx", excel_bytes,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("initiate_shipment_data", (None, initiate_payload.encode("utf-8"),
                                        "application/json")),
        ]

        async with httpx.AsyncClient(timeout=90) as client:
            init_resp = await client.post(
                f"{base_url}/initiate", files=initiate_files, headers=headers,
            )

        if init_resp.status_code == 401 and attempt == 0:
            logger.warning("Xindus initiate returned 401, refreshing token")
            _clear_token()
            continue

        try:
            init_body = init_resp.json()
        except Exception:
            init_body = {"raw_response": init_resp.text[:2000]}

        if init_resp.status_code != 200:
            logger.error("Xindus initiate failed: %d %s",
                         init_resp.status_code, str(init_body)[:300])
            return init_resp.status_code, init_body

        # Extract scancode from initiate response
        # Response shape: {"data": [{"scancode": "DR000000XXX", ...}]}
        init_data = init_body.get("data", [{}])
        if isinstance(init_data, list) and init_data:
            scancode = init_data[0].get("scancode")
        else:
            scancode = None

        if not scancode:
            logger.error("No scancode in initiate response: %s", init_body)
            return 500, {"error": "Initiate succeeded but no scancode returned",
                         "initiate_response": init_body}

        logger.info("Xindus initiate OK, scancode=%s", scancode)

        # ── Step 2: Create — update the initiated shipment ──
        create_data = {**snake_data, "scancode": scancode}
        create_payload = json.dumps(create_data)
        create_files = [
            ("box_details_file", ("uploadedFile.xlsx", excel_bytes,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("create_shipment_data", (None, create_payload.encode("utf-8"),
                                      "application/json")),
        ]

        async with httpx.AsyncClient(timeout=90) as client:
            resp = await client.post(
                f"{base_url}/create", files=create_files, headers=headers,
            )

        if resp.status_code == 401 and attempt == 0:
            logger.warning("Xindus create returned 401, refreshing token")
            _clear_token()
            continue

        try:
            body = resp.json()
        except Exception:
            body = {"raw_response": resp.text[:2000]}

        logger.info("Xindus express-shipment create response: %d %s",
                     resp.status_code, str(body)[:300])
        return resp.status_code, body

    return 500, {"error": "Failed after retry"}
