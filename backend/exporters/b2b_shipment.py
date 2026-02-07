"""B2B Shipment block-format Excel exporter.

Generates a single-sheet workbook using the repeating-block layout where
each shipment is a vertical block:

    Row 1: Address header (#record_type, #receiver_name, ...)
    Row 2: Address data   (address, Amazon FBA FTW1, ...)
    Row 3: Box/item header (#box_number, Box Length, ...)
    Row 4+: Box/item data  (1, 5, 20, 10, 1, Utensils, ...)

Multiple shipments stack vertically with no blank rows between blocks.
"""

from __future__ import annotations

from difflib import SequenceMatcher
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.models.extraction import (
    Address,
    Box,
    BoxItem,
    ExtractionResult,
    LineItem,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Address header row columns (A-J)
ADDRESS_HEADERS = [
    "#record_type",
    "#receiver_name(R)",
    "#receiver_address(R)",
    "#receiver_city(R)",
    "#receiver_zip(R)",
    "#receiver_state(R)",
    "#receiver_country(R)",
    "#receiver_phone_number(R)",
    "#receiver_extension_no(O)",
    "#receiver_email(R)",
]

# Box/item header row columns (A-L)
BOX_ITEM_HEADERS = [
    "#box_number(R)",
    "Box Length(cms)(R)",
    "Box Width(cms)(R)",
    "Box Height(cms)(R)",
    "Box Weight(kgs)(R)",
    "Item Description(R)",
    "Item Qty(R)",
    "Item Unit Weight Kg(O)",
    "HS Code(Origin)(R)",
    "HS Code(Destination)(O)",
    "Item Unit Price(R)",
    "IGST % (For GST taxtype)",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _v(cv) -> Optional[str | int | float]:
    """Extract the raw value from a ConfidenceValue."""
    return cv.value if cv else None


def _address_is_empty(addr: Address) -> bool:
    return all(
        _v(f) in (None, "")
        for f in [addr.name, addr.address, addr.city, addr.state,
                  addr.zip_code, addr.country, addr.phone]
    )


def _similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _match_box_item_to_line_item(
    box_item: BoxItem,
    line_items: list[LineItem],
    threshold: float = 0.4,
) -> Optional[LineItem]:
    box_desc = str(_v(box_item.description) or "")
    if not box_desc:
        return None

    best_match: Optional[LineItem] = None
    best_score = 0.0

    for li in line_items:
        li_desc = str(_v(li.description) or "")
        score = _similarity(box_desc, li_desc)
        if score > best_score:
            best_score = score
            best_match = li

    return best_match if best_score >= threshold else None


def _build_item_rows_for_box(
    box: Box,
    line_items: list[LineItem],
) -> list[dict]:
    rows: list[dict] = []

    if box.items:
        for bi in box.items:
            matched = _match_box_item_to_line_item(bi, line_items)
            row = {
                "description": _v(bi.description) or (_v(matched.description) if matched else ""),
                "quantity": _v(bi.quantity) or (_v(matched.quantity) if matched else ""),
                "unit_weight": _v(matched.unit_weight_kg) if matched else "",
                "hs_origin": _v(matched.hs_code_origin) if matched else "",
                "hs_dest": _v(matched.hs_code_destination) if matched else "",
                "unit_price": _v(matched.unit_price_usd) if matched else "",
                "igst": _v(matched.igst_percent) if matched else "",
            }
            rows.append(row)
    else:
        for li in line_items:
            row = {
                "description": _v(li.description) or "",
                "quantity": _v(li.quantity) or "",
                "unit_weight": _v(li.unit_weight_kg) or "",
                "hs_origin": _v(li.hs_code_origin) or "",
                "hs_dest": _v(li.hs_code_destination) or "",
                "unit_price": _v(li.unit_price_usd) or "",
                "igst": _v(li.igst_percent) or "",
            }
            rows.append(row)

    if not rows:
        rows.append({
            "description": "", "quantity": "", "unit_weight": "",
            "hs_origin": "", "hs_dest": "", "unit_price": "", "igst": "",
        })

    return rows


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_data_row(ws, row: int, values: list, max_col: int = 12) -> None:
    """Write a data row with border and font."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value if value not in (None, "") else "")
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # Fill remaining columns up to max_col with bordered empty cells
    for col_idx in range(len(values) + 1, max_col + 1):
        cell = ws.cell(row=row, column=col_idx, value="")
        cell.font = DATA_FONT
        cell.border = THIN_BORDER


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 35) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _address_hash(addr: Address) -> str:
    """Create a grouping key from an address (name+city+zip)."""
    parts = [
        str(_v(addr.name) or "").lower().strip(),
        str(_v(addr.city) or "").lower().strip(),
        str(_v(addr.zip_code) or "").lower().strip(),
    ]
    return "|".join(parts)


def _group_boxes_by_receiver(
    boxes: list[Box], global_receiver: Address
) -> list[tuple[Address, list[Box]]]:
    """Group boxes by their resolved receiver address.

    Returns a list of (receiver_address, boxes_for_that_receiver) tuples,
    preserving the order in which each receiver was first encountered.
    """
    groups: dict[str, tuple[Address, list[Box]]] = {}
    order: list[str] = []

    for box in boxes:
        receiver = box.receiver if (box.receiver and not _address_is_empty(box.receiver)) else global_receiver
        key = _address_hash(receiver)
        if key not in groups:
            groups[key] = (receiver, [])
            order.append(key)
        groups[key][1].append(box)

    return [groups[k] for k in order]


def generate_b2b_shipment(result: ExtractionResult, output_path: str) -> str:
    """Generate a B2B Shipment block-format Excel file.

    Groups boxes by receiver address and writes one shipment block per
    destination group (address header -> address data -> box header -> box rows).

    Parameters
    ----------
    result : ExtractionResult
        Structured extraction data from the document processing pipeline.
    output_path : str
        Destination file path for the generated .xlsx.

    Returns
    -------
    str
        The *output_path* that was written.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "B2B Shipment"

    # ---- Determine global fallback receiver ----
    ship_to = result.invoice.ship_to
    global_receiver = ship_to if not _address_is_empty(ship_to) else result.invoice.consignee

    line_items = result.invoice.line_items
    boxes = result.packing_list.boxes

    # If there are no boxes, create a single virtual box
    if not boxes:
        virtual_box = Box()
        virtual_box.box_number.value = 1
        boxes = [virtual_box]

    # Group boxes by destination
    receiver_groups = _group_boxes_by_receiver(boxes, global_receiver)

    current_row = 1

    for receiver, group_boxes in receiver_groups:
        # Section 1: Address header
        _write_header_row(ws, current_row, ADDRESS_HEADERS)
        current_row += 1

        # Section 1: Address data
        address_data = [
            "address",
            _v(receiver.name) or "",
            _v(receiver.address) or "",
            _v(receiver.city) or "",
            _v(receiver.zip_code) or "",
            _v(receiver.state) or "",
            _v(receiver.country) or "",
            _v(receiver.phone) or "",
            "",  # Extension -- not in extraction model
            _v(receiver.email) or "",
        ]
        _write_data_row(ws, current_row, address_data)
        current_row += 1

        # Section 2: Box/item header
        _write_header_row(ws, current_row, BOX_ITEM_HEADERS)
        current_row += 1

        # Section 2: Box/item data rows
        for box in group_boxes:
            item_rows = _build_item_rows_for_box(box, line_items)

            for item in item_rows:
                box_item_data = [
                    _v(box.box_number) or "",
                    _v(box.length_cm) or "",
                    _v(box.width_cm) or "",
                    _v(box.height_cm) or "",
                    _v(box.gross_weight_kg) or "",
                    item["description"],
                    item["quantity"],
                    item["unit_weight"],
                    item["hs_origin"],
                    item["hs_dest"],
                    item["unit_price"],
                    item["igst"],
                ]
                _write_data_row(ws, current_row, box_item_data)
                current_row += 1

    # ---- Finishing touches ----
    _auto_fit_columns(ws)

    wb.save(output_path)
    return output_path
