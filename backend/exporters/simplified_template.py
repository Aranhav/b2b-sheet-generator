"""Simplified hierarchical 4-sheet Excel exporter.

Generates a workbook with four sheets:
  1. Items      -- one row per unique invoice line item
  2. Receivers  -- one row per unique destination address
  3. Boxes      -- one row per item-in-box with references to Items and Receivers
  4. Standard Addresses -- pre-populated Amazon FBA & Walmart WFS addresses
"""

from __future__ import annotations

from difflib import SequenceMatcher
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.models.extraction import (
    Address,
    Box,
    BoxItem,
    ExtractionResult,
    LineItem,
)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Items / Boxes title & header
BLUE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Receivers / Standard Addresses header
GREEN_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

# Boxes header
GOLD_FILL = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
BLUE_BOLD = Font(bold=True, color="2F5496", size=13)
ITALIC_FONT = Font(italic=True, size=10, color="333333")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Tab colours (hex without #)
TAB_ITEMS = "2F5496"
TAB_RECEIVERS = "548235"
TAB_BOXES = "BF8F00"
TAB_STD_ADDRESSES = "548235"

# Column width bounds
MIN_COL_WIDTH = 12
MAX_COL_WIDTH = 35

# ---------------------------------------------------------------------------
# Standard FBA / WFS addresses
# ---------------------------------------------------------------------------

_STANDARD_ADDRESSES: list[dict[str, str]] = [
    # Amazon FBA US
    {"id": "FBA-FTW1", "type": "Amazon FBA", "name": "Amazon.com Services LLC - FTW1", "address": "33333 Lyndon B Johnson Fwy", "city": "Dallas", "state": "TX", "zip": "75241", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-ONT8", "type": "Amazon FBA", "name": "Amazon.com Services LLC - ONT8", "address": "24300 Nandina Ave", "city": "Moreno Valley", "state": "CA", "zip": "92551", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-BOS7", "type": "Amazon FBA", "name": "Amazon.com Services LLC - BOS7", "address": "1000 Technology Center Dr", "city": "Fall River", "state": "MA", "zip": "02723", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-PHX6", "type": "Amazon FBA", "name": "Amazon.com Services LLC - PHX6", "address": "4750 W Mohave St", "city": "Phoenix", "state": "AZ", "zip": "85043", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-MCO1", "type": "Amazon FBA", "name": "Amazon.com Services LLC - MCO1", "address": "2850 Penny Rd", "city": "Apopka", "state": "FL", "zip": "32703", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-JFK8", "type": "Amazon FBA", "name": "Amazon.com Services LLC - JFK8", "address": "546 Gulf Ave", "city": "Staten Island", "state": "NY", "zip": "10314", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-MIA1", "type": "Amazon FBA", "name": "Amazon.com Services LLC - MIA1", "address": "14000 NW 37th Ave", "city": "Opa-locka", "state": "FL", "zip": "33054", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-IND1", "type": "Amazon FBA", "name": "Amazon.com Services LLC - IND1", "address": "4255 Anson Blvd", "city": "Whitestown", "state": "IN", "zip": "46075", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-SDF8", "type": "Amazon FBA", "name": "Amazon.com Services LLC - SDF8", "address": "900 Patrol Rd", "city": "Jeffersonville", "state": "IN", "zip": "47130", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-DFW7", "type": "Amazon FBA", "name": "Amazon.com Services LLC - DFW7", "address": "700 Westport Pkwy", "city": "Fort Worth", "state": "TX", "zip": "76177", "country": "US", "phone": "", "email": ""},
    {"id": "FBA-EWR4", "type": "Amazon FBA", "name": "Amazon.com Services LLC - EWR4", "address": "50 New Canton Way", "city": "Robbinsville", "state": "NJ", "zip": "08691", "country": "US", "phone": "", "email": ""},
    # Walmart WFS
    {"id": "WFS-LAX1", "type": "Walmart WFS", "name": "Walmart Fulfillment Services - LAX1", "address": "17067 Edison Ave", "city": "Chino", "state": "CA", "zip": "91708", "country": "US", "phone": "", "email": ""},
    {"id": "WFS-IND1", "type": "Walmart WFS", "name": "Walmart Fulfillment Services - IND1", "address": "8401 Bearing Dr", "city": "Indianapolis", "state": "IN", "zip": "46268", "country": "US", "phone": "", "email": ""},
    {"id": "WFS-ATL1", "type": "Walmart WFS", "name": "Walmart Fulfillment Services - ATL1", "address": "615 Ga Hwy 18 Connector", "city": "West Point", "state": "GA", "zip": "31833", "country": "US", "phone": "", "email": ""},
    {"id": "WFS-NJ3", "type": "Walmart WFS", "name": "Walmart Fulfillment Services - NJ3", "address": "690 Newark Turnpike", "city": "Kearny", "state": "NJ", "zip": "07032", "country": "US", "phone": "", "email": ""},
    {"id": "WFS-PHX1", "type": "Walmart WFS", "name": "Walmart Fulfillment Services - PHX1", "address": "5765 S Sossaman Rd", "city": "Mesa", "state": "AZ", "zip": "85212", "country": "US", "phone": "", "email": ""},
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _v(cv) -> Optional[str | int | float]:
    """Extract the raw value from a ConfidenceValue."""
    return cv.value if cv else None


def _address_is_empty(addr: Address) -> bool:
    return all(
        _v(f) in (None, "")
        for f in [addr.name, addr.address, addr.city, addr.state, addr.zip_code, addr.country, addr.phone]
    )


def _similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _match_box_item_to_line_items(
    box_item: BoxItem,
    line_items: list[LineItem],
    threshold: float = 0.4,
) -> Optional[int]:
    """Return the 0-based index of the best-matching line item, or None."""
    box_desc = str(_v(box_item.description) or "")
    if not box_desc:
        return None

    best_idx: Optional[int] = None
    best_score = 0.0

    for idx, li in enumerate(line_items):
        li_desc = str(_v(li.description) or "")
        score = _similarity(box_desc, li_desc)
        if score > best_score:
            best_score = score
            best_idx = idx

    return best_idx if best_score >= threshold else None


def _auto_fit(ws, min_width: int = MIN_COL_WIDTH, max_width: int = MAX_COL_WIDTH) -> None:
    """Set column widths based on content, clamped to [min_width, max_width]."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def _apply_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Apply thin borders to a rectangular region."""
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_items_sheet(ws, result: ExtractionResult) -> None:
    """Sheet 1: Items -- one row per unique invoice line item."""
    ws.title = "Items"
    ws.sheet_properties.tabColor = TAB_ITEMS

    headers = [
        "Item ID*",
        "Item Description*",
        "HS Code (Origin)*",
        "HS Code (Destination)",
        "Item Unit Price (USD)*",
        "Item Unit Weight (Kg)",
        "IGST %",
    ]
    descriptions = [
        "Auto-generated ID",
        "Full product description",
        "HS/HTS code in origin country",
        "HS/HTS code in destination country",
        "Price per unit in USD",
        "Weight per unit in kilograms",
        "IGST percentage if applicable",
    ]
    # Required column indices (1-based): 1, 2, 3, 5
    required_cols = {1, 2, 3, 5}

    col_count = len(headers)

    # Row 1 -- title (merged across all columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value="ITEMS \u2014 Enter each unique product ONCE")
    title_cell.font = BLUE_BOLD
    title_cell.alignment = LEFT_ALIGN

    # Row 2 -- column descriptions
    for col_idx, desc in enumerate(descriptions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.fill = LIGHT_BLUE_FILL
        cell.font = ITALIC_FONT
        cell.alignment = LEFT_ALIGN

    # Row 3 -- headers
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.fill = BLUE_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Row 4+ -- data
    for i, li in enumerate(result.invoice.line_items):
        row_num = 4 + i
        item_id = f"I{i + 1}"
        values = [
            item_id,
            _v(li.description) or "",
            _v(li.hs_code_origin) or "",
            _v(li.hs_code_destination) or "",
            _v(li.unit_price_usd) or "",
            _v(li.unit_weight_kg) or "",
            _v(li.igst_percent) or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col_idx in required_cols:
                cell.fill = YELLOW_FILL

    # Borders on description row
    _apply_border(ws, 2, 2, 1, col_count)

    # Freeze after header row (row 3)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(col_count)}3"
    _auto_fit(ws)


def _build_receivers_sheet(ws, result: ExtractionResult) -> list[tuple[str, Address]]:
    """Sheet 2: Receivers -- one row per unique destination.

    Uses packing_list.destinations when available (multi-address extraction),
    falling back to ship_to/consignee/ior for legacy single-address results.

    Returns a list of (receiver_id, Address) for use by the Boxes sheet.
    """
    ws.title = "Receivers"
    ws.sheet_properties.tabColor = TAB_RECEIVERS

    headers = [
        "Receiver ID*",
        "Receiver Name*",
        "Receiver Address*",
        "Receiver City*",
        "Receiver Zip*",
        "Receiver State*",
        "Receiver Country*",
        "Receiver Phone*",
        "Receiver Extension",
        "Receiver Email*",
    ]
    required_cols = {1, 2, 3, 4, 5, 6, 7, 8, 10}

    col_count = len(headers)

    # Row 1 -- title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value="RECEIVERS")
    title_cell.font = BLUE_BOLD
    title_cell.alignment = LEFT_ALIGN

    # Row 2 -- blank spacer (keeps structure consistent with Items)
    # Row 3 -- headers
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.fill = GREEN_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Determine receivers -- prefer destinations from packing list extraction
    receivers: list[tuple[str, Address]] = []

    destinations = result.packing_list.destinations
    if destinations:
        for dest in destinations:
            rid = f"R-{dest.id}" if dest.id else f"R{len(receivers) + 1}"
            receivers.append((rid, dest.address))
    else:
        # Fallback: legacy single-address mode
        ship_to = result.invoice.ship_to
        consignee = result.invoice.consignee
        primary = ship_to if not _address_is_empty(ship_to) else consignee
        if not _address_is_empty(primary):
            receivers.append(("R1", primary))

    ior = result.invoice.ior
    if not _address_is_empty(ior):
        receivers.append((f"R-IOR", ior))

    # Row 4+ -- data
    for i, (rid, addr) in enumerate(receivers):
        row_num = 4 + i
        values = [
            rid,
            _v(addr.name) or "",
            _v(addr.address) or "",
            _v(addr.city) or "",
            _v(addr.zip_code) or "",
            _v(addr.state) or "",
            _v(addr.country) or "",
            _v(addr.phone) or "",
            "",  # Extension -- not in model
            _v(addr.email) or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col_idx in required_cols:
                cell.fill = YELLOW_FILL

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(col_count)}3"
    _auto_fit(ws)

    return receivers


def _build_boxes_sheet(ws, result: ExtractionResult) -> None:
    """Sheet 3: Boxes -- one row per item-in-box."""
    ws.title = "Boxes"
    ws.sheet_properties.tabColor = TAB_BOXES

    headers = [
        "Box Number*",
        "Receiver ID*",
        "Box Length (cms)*",
        "Box Width (cms)*",
        "Box Height (cms)*",
        "Box Weight (kgs)*",
        "Item ID*",
        "Item Qty*",
    ]
    required_cols = set(range(1, len(headers) + 1))  # all required

    col_count = len(headers)

    # Row 1 -- title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value="BOXES")
    title_cell.font = BLUE_BOLD
    title_cell.alignment = LEFT_ALIGN

    # Row 3 -- headers
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.fill = GOLD_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    line_items = result.invoice.line_items
    boxes = result.packing_list.boxes
    destinations = result.packing_list.destinations

    # Build destination_id -> receiver_id mapping
    dest_to_receiver: dict[str, str] = {}
    for dest in destinations:
        dest_to_receiver[dest.id] = f"R-{dest.id}" if dest.id else "R1"

    # Build item ID lookup from line item descriptions
    item_id_map: dict[str, str] = {}
    for idx, li in enumerate(line_items):
        desc = str(_v(li.description) or "").lower().strip()
        if desc:
            item_id_map[desc] = f"I{idx + 1}"

    current_row = 4

    for box in boxes:
        box_num = _v(box.box_number) or ""

        # Determine receiver ID for this box
        dest_id = _v(box.destination_id)
        receiver_id = dest_to_receiver.get(str(dest_id), "R1") if dest_id else "R1"

        if box.items:
            # Box has explicit items
            for item_idx, bi in enumerate(box.items):
                is_first = item_idx == 0

                # Try to match to a line item for the Item ID
                matched_idx = _match_box_item_to_line_items(bi, line_items)
                item_id = f"I{matched_idx + 1}" if matched_idx is not None else ""
                qty = _v(bi.quantity) or ""

                if is_first:
                    values = [
                        box_num,
                        receiver_id,
                        _v(box.length_cm) or "",
                        _v(box.width_cm) or "",
                        _v(box.height_cm) or "",
                        _v(box.gross_weight_kg) or "",
                        item_id,
                        qty,
                    ]
                else:
                    values = [
                        "",  # Box number blank for subsequent items
                        "",  # Receiver ID blank
                        "",  # Length blank
                        "",  # Width blank
                        "",  # Height blank
                        "",  # Weight blank
                        item_id,
                        qty,
                    ]

                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.border = THIN_BORDER
                    cell.alignment = LEFT_ALIGN
                    if col_idx in required_cols and val not in ("", None):
                        cell.fill = YELLOW_FILL

                current_row += 1
        else:
            # No explicit items -- distribute all invoice items into this box
            for item_idx, li in enumerate(line_items):
                is_first = item_idx == 0
                item_id = f"I{item_idx + 1}"
                qty = _v(li.quantity) or ""

                if is_first:
                    values = [
                        box_num,
                        receiver_id,
                        _v(box.length_cm) or "",
                        _v(box.width_cm) or "",
                        _v(box.height_cm) or "",
                        _v(box.gross_weight_kg) or "",
                        item_id,
                        qty,
                    ]
                else:
                    values = [
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        item_id,
                        qty,
                    ]

                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.border = THIN_BORDER
                    cell.alignment = LEFT_ALIGN
                    if col_idx in required_cols and val not in ("", None):
                        cell.fill = YELLOW_FILL

                current_row += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(col_count)}3"
    _auto_fit(ws)


def _build_standard_addresses_sheet(ws) -> None:
    """Sheet 4: Standard Addresses -- pre-populated FBA & WFS addresses."""
    ws.title = "Standard Addresses"
    ws.sheet_properties.tabColor = TAB_STD_ADDRESSES

    headers = [
        "Address ID",
        "Type",
        "Name",
        "Address",
        "City",
        "State",
        "ZIP",
        "Country",
        "Phone",
        "Email",
    ]
    col_count = len(headers)

    # Row 1 -- title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value="STANDARD ADDRESSES \u2014 Reference Only")
    title_cell.font = BLUE_BOLD
    title_cell.alignment = LEFT_ALIGN

    # Row 2 -- blank spacer
    # Row 3 -- headers
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.fill = GREEN_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Row 4+ -- data
    for i, entry in enumerate(_STANDARD_ADDRESSES):
        row_num = 4 + i
        values = [
            entry["id"],
            entry["type"],
            entry["name"],
            entry["address"],
            entry["city"],
            entry["state"],
            entry["zip"],
            entry["country"],
            entry["phone"],
            entry["email"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(col_count)}{3 + len(_STANDARD_ADDRESSES)}"
    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_simplified_template(result: ExtractionResult, output_path: str) -> str:
    """Generate a 4-sheet simplified-template Excel file.

    Parameters
    ----------
    result : ExtractionResult
        Structured extraction data from the document processing pipeline.
    output_path : str
        Destination file path for the generated .xlsx.

    Returns
    -------
    str
        The *output_path* that was written.
    """
    wb = Workbook()

    # Sheet 1 -- Items (uses the default sheet created by Workbook())
    ws_items = wb.active
    _build_items_sheet(ws_items, result)

    # Sheet 2 -- Receivers
    ws_receivers = wb.create_sheet()
    _build_receivers_sheet(ws_receivers, result)

    # Sheet 3 -- Boxes
    ws_boxes = wb.create_sheet()
    _build_boxes_sheet(ws_boxes, result)

    # Sheet 4 -- Standard Addresses
    ws_std = wb.create_sheet()
    _build_standard_addresses_sheet(ws_std)

    wb.save(output_path)
    return output_path
