"""XpressB2B Multi Address flat-format Excel exporter.

Generates a single-sheet workbook with one row per item-in-box,
following the legacy XpressB2B bulk-upload column layout (21 columns).
"""

from __future__ import annotations

from difflib import SequenceMatcher
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.models.extraction import (
    Address,
    Box,
    BoxItem,
    ExtractionResult,
    LineItem,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMNS = [
    "Box Number",
    "Receiver Name",
    "Receiver Address",
    "Receiver City",
    "Receiver Zip",
    "Receiver State",
    "Receiver Country",
    "Receiver Phone Number",
    "Receiver Extension No",
    "Receiver Email",
    "Box Length (cms)",
    "Box Width (cms)",
    "Box Height (cms)",
    "Box Weight (kgs)",
    "Item Description",
    "Item Qty",
    "Item Unit Weight Kg",
    "HS Code (Origin)",
    "HS Code (Destination)",
    "Item Unit Price",
    "IGST % (For GST taxtype)",
]

# Number of "box-level" columns that are only filled on the first row of a box.
BOX_LEVEL_COL_COUNT = 14


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _v(cv) -> Optional[str | int | float]:
    """Extract the raw value from a ConfidenceValue."""
    return cv.value if cv else None


def _address_is_empty(addr: Address) -> bool:
    """Return True if all significant address fields are empty."""
    return all(
        _v(f) in (None, "")
        for f in [addr.name, addr.address, addr.city, addr.state, addr.zip_code, addr.country, addr.phone]
    )


def _similarity(a: str, b: str) -> float:
    """Return a 0..1 similarity ratio between two strings (case-insensitive)."""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _match_box_item_to_line_item(
    box_item: BoxItem,
    line_items: list[LineItem],
    threshold: float = 0.4,
) -> Optional[LineItem]:
    """Find the best-matching invoice line item for a given box item."""
    box_desc = str(_v(box_item.description) or "")
    if not box_desc:
        return None

    best_match: Optional[LineItem] = None
    best_score = 0.0

    for li in line_items:
        li_desc = str(_v(li.description) or "")
        score = _similarity(box_desc, li_desc)
        if score > best_score:
            best_score = score
            best_match = li

    return best_match if best_score >= threshold else None


def _build_item_rows_for_box(
    box: Box,
    line_items: list[LineItem],
) -> list[dict]:
    """Build one dict per item row for a given box.

    If the box has explicit items, match each to an invoice line item.
    If the box has no items, distribute all invoice line items across it.
    """
    rows: list[dict] = []

    if box.items:
        for bi in box.items:
            matched = _match_box_item_to_line_item(bi, line_items)
            row = {
                "description": _v(bi.description) or (_v(matched.description) if matched else ""),
                "quantity": _v(bi.quantity) or (_v(matched.quantity) if matched else ""),
                "unit_weight": _v(matched.unit_weight_kg) if matched else "",
                "hs_origin": _v(matched.hs_code_origin) if matched else "",
                "hs_dest": _v(matched.hs_code_destination) if matched else "",
                "unit_price": _v(matched.unit_price_usd) if matched else "",
                "igst": _v(matched.igst_percent) if matched else "",
            }
            rows.append(row)
    else:
        # No explicit box items -- put all invoice line items in this box
        for li in line_items:
            row = {
                "description": _v(li.description) or "",
                "quantity": _v(li.quantity) or "",
                "unit_weight": _v(li.unit_weight_kg) or "",
                "hs_origin": _v(li.hs_code_origin) or "",
                "hs_dest": _v(li.hs_code_destination) or "",
                "unit_price": _v(li.unit_price_usd) or "",
                "igst": _v(li.igst_percent) or "",
            }
            rows.append(row)

    # Guarantee at least one row so the box dimensions still appear
    if not rows:
        rows.append({
            "description": "",
            "quantity": "",
            "unit_weight": "",
            "hs_origin": "",
            "hs_dest": "",
            "unit_price": "",
            "igst": "",
        })

    return rows


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Adjust column widths based on content length."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_multi_address(result: ExtractionResult, output_path: str) -> str:
    """Generate an XpressB2B Multi-Address Excel file.

    Parameters
    ----------
    result : ExtractionResult
        Structured extraction data from the document processing pipeline.
    output_path : str
        Destination file path for the generated .xlsx.

    Returns
    -------
    str
        The *output_path* that was written.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "XpressB2B Multi Address"

    # ---- Header row ----
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # ---- Determine global fallback receiver address ----
    ship_to = result.invoice.ship_to
    global_receiver = ship_to if not _address_is_empty(ship_to) else result.invoice.consignee

    line_items = result.invoice.line_items
    boxes = result.packing_list.boxes

    # If there are no boxes at all, create a single virtual box
    if not boxes:
        virtual_box = Box()
        virtual_box.box_number.value = 1
        boxes = [virtual_box]

    # ---- Data rows ----
    current_row = 2

    for box in boxes:
        item_rows = _build_item_rows_for_box(box, line_items)

        # Per-box receiver: use box.receiver if populated, else global fallback
        receiver = box.receiver if (box.receiver and not _address_is_empty(box.receiver)) else global_receiver

        for item_idx, item in enumerate(item_rows):
            is_first = item_idx == 0

            if is_first:
                row_data = [
                    _v(box.box_number) or "",
                    _v(receiver.name) or "",
                    _v(receiver.address) or "",
                    _v(receiver.city) or "",
                    _v(receiver.zip_code) or "",
                    _v(receiver.state) or "",
                    _v(receiver.country) or "",
                    _v(receiver.phone) or "",
                    "",  # Extension No -- not in extraction model
                    _v(receiver.email) or "",
                    _v(box.length_cm) or "",
                    _v(box.width_cm) or "",
                    _v(box.height_cm) or "",
                    _v(box.gross_weight_kg) or "",
                ]
            else:
                # Subsequent items in the same box: box-level columns blank
                row_data = [""] * BOX_LEVEL_COL_COUNT

            # Item-level columns (15-21)
            row_data.extend([
                item["description"],
                item["quantity"],
                item["unit_weight"],
                item["hs_origin"],
                item["hs_dest"],
                item["unit_price"],
                item["igst"],
            ])

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            current_row += 1

    # ---- Finishing touches ----
    _auto_fit_columns(ws)
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path
