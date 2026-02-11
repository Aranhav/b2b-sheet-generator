"""B2B Booking Agent API router.

Endpoints:
  POST   /api/agent/upload            Upload files, create batch, start extraction
  GET    /api/agent/jobs/{id}/stream   SSE progress stream
  GET    /api/agent/batches/{id}       Get batch with all drafts
  GET    /api/agent/drafts/{id}        Get single draft details
  PATCH  /api/agent/drafts/{id}        Apply corrections
  POST   /api/agent/drafts/{id}/approve  Push to Xindus API
  POST   /api/agent/drafts/{id}/archive Archive a draft
  POST   /api/agent/drafts/{id}/delete  Permanently delete a draft
  POST   /api/agent/drafts/{id}/re-extract  Re-run extraction on draft files
  POST   /api/agent/drafts/{id}/classify   Run Gaia tariff classification
  POST   /api/agent/tariff-lookup          Tariff detail lookup for a single HTS code
  DELETE /api/agent/drafts/{id}        Soft-delete (reject) draft
"""
from __future__ import annotations

import asyncio
import json
import time
import logging
from typing import Any
from uuid import UUID

import httpx
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import Response as FastAPIResponse
from sse_starlette.sse import EventSourceResponse

from backend import db, storage
from backend.config import ALLOWED_EXTENSIONS, MAX_FILE_SIZE_MB, normalize_country_code
from backend.models.agent import (
    ActiveBatch,
    ActiveBatchesResponse,
    ApprovalResponse,
    BatchResponse,
    CorrectionRequest,
    DraftDetail,
    DraftsListResponse,
    DraftSummary,
    FileInfo,
    SellerProfile,
    SellersListResponse,
    SSEProgress,
    SubmissionResult,
    SubmitToXindusRequest,
    UploadResponse,
    XindusAddress,
    XindusCustomer,
    XindusSyncRequest,
    XindusSyncResponse,
)
from backend.services.classifier import classify_document
from backend.services.draft_builder import build_draft_shipment
from backend.services.grouper import group_files_into_shipments
from backend.services.pdf_processor import process_pdf

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/agent", tags=["agent"])

# ---------------------------------------------------------------------------
# In-memory job progress tracking (lightweight; no Redis needed)
# ---------------------------------------------------------------------------

_job_queues: dict[str, asyncio.Queue] = {}


async def _send_progress(job_id: str, batch_id: UUID, event: SSEProgress) -> None:
    """Push an SSE event to the job's queue and persist step to DB."""
    q = _job_queues.get(job_id)
    if q:
        q.put_nowait(event)

    # Persist step + progress to DB for recovery after disconnect
    if event.step not in ("complete", "error"):
        progress_data: dict[str, Any] = {
            "completed": event.completed,
            "total": event.total,
        }
        if event.file:
            progress_data["file"] = event.file
        if event.shipments_found is not None:
            progress_data["shipments_found"] = event.shipments_found
        try:
            await db.update_batch_progress(batch_id, event.step, progress_data)
        except Exception:
            logger.warning("Failed to persist batch progress", exc_info=True)


# ---------------------------------------------------------------------------
# POST /api/agent/upload
# ---------------------------------------------------------------------------


@router.post("/upload", response_model=UploadResponse)
async def upload_files(files: list[UploadFile] = File(...)):
    """Accept file uploads, create a batch, and start async extraction."""
    if not files:
        raise HTTPException(400, "No files uploaded")

    # Validate files
    for f in files:
        ext = "." + f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(400, f"Unsupported file type: {f.filename}")

    # Create batch
    batch_id = await db.create_batch(file_count=len(files))
    job_id = str(batch_id)

    # Create progress queue
    _job_queues[job_id] = asyncio.Queue()

    # Read file bytes and store records
    file_records: list[dict[str, Any]] = []
    for f in files:
        content = await f.read()

        # Check file size
        if len(content) > MAX_FILE_SIZE_MB * 1024 * 1024:
            raise HTTPException(400, f"File too large: {f.filename} (max {MAX_FILE_SIZE_MB}MB)")

        # Store in S3 / local
        s3_key = f"uploads/{batch_id}/{f.filename}"
        file_url = await storage.upload_file(s3_key, content)

        # Create DB record
        file_id = await db.create_file(batch_id, f.filename, file_url)
        file_records.append({
            "id": file_id,
            "filename": f.filename,
            "content": content,
        })

    # Start extraction in background
    asyncio.create_task(_run_extraction_pipeline(job_id, batch_id, file_records))

    return UploadResponse(batch_id=batch_id, file_count=len(files))


# ---------------------------------------------------------------------------
# Extraction pipeline (runs in background)
# ---------------------------------------------------------------------------


async def _run_extraction_pipeline(
    job_id: str,
    batch_id: UUID,
    file_records: list[dict[str, Any]],
) -> None:
    """Full extraction pipeline: classify → extract → group → build drafts."""
    total = len(file_records)

    try:
        # ── Step 1: Process PDFs and classify ──────────────────────
        extracted_files: list[dict[str, Any]] = []

        for idx, fr in enumerate(file_records):
            await _send_progress(job_id, batch_id, SSEProgress(
                step="classifying",
                file=fr["filename"],
                completed=idx,
                total=total,
            ))

            # Process PDF (text extraction / image rendering)
            pdf_result = await asyncio.to_thread(process_pdf, fr["content"])
            pages = pdf_result.get("pages", [])
            is_vision = any(p.get("image_bytes") for p in pages)

            # Classify document type
            classification = await classify_document(pages, is_vision=is_vision)
            file_type = classification["document_type"]

            await db.update_file_extraction(
                fr["id"],
                file_type=file_type,
                page_count=len(pages),
            )

            extracted_files.append({
                "id": fr["id"],
                "filename": fr["filename"],
                "file_type": file_type,
                "pages": pages,
                "is_vision": is_vision,
                "classification": classification,
            })

        await _send_progress(job_id, batch_id, SSEProgress(
            step="classifying",
            completed=total,
            total=total,
        ))

        # ── Step 2: Extract data per file ──────────────────────────
        from backend.services.llm_extractor import _extract_invoice, _extract_packing_list
        from backend.services.learning import build_extraction_hint

        # Build correction hints from past user corrections (self-improvement)
        try:
            correction_hints = await build_extraction_hint()
        except Exception:
            logger.warning("Failed to build extraction hints, continuing without", exc_info=True)
            correction_hints = ""

        for idx, ef in enumerate(extracted_files):
            await _send_progress(job_id, batch_id, SSEProgress(
                step="extracting",
                file=ef["filename"],
                completed=idx,
                total=total,
            ))

            pages = ef["pages"]
            is_vision = ef["is_vision"]
            file_type = ef["file_type"]

            try:
                if file_type == "invoice":
                    raw_data = await _extract_invoice(pages, is_vision, correction_hints=correction_hints)
                elif file_type == "packing_list":
                    raw_data = await _extract_packing_list(pages, is_vision, correction_hints=correction_hints)
                else:
                    # For certificates and other types, try invoice extraction
                    # to capture any identifiable metadata
                    raw_data = await _extract_invoice(pages, is_vision, correction_hints=correction_hints)
            except Exception:
                logger.exception("Extraction failed for %s", ef["filename"])
                raw_data = {}

            ef["extracted_data"] = raw_data

            # Compute a simple confidence from extracted data
            confidence = _estimate_confidence(raw_data, file_type)

            await db.update_file_extraction(
                ef["id"],
                extracted_data=raw_data,
                confidence=confidence,
            )

        await _send_progress(job_id, batch_id, SSEProgress(
            step="extracting",
            completed=total,
            total=total,
        ))

        # ── Step 3: Group files into shipments ─────────────────────
        await _send_progress(job_id, batch_id, SSEProgress(step="grouping", total=total))

        group_input = [
            {
                "id": ef["id"],
                "file_type": ef["file_type"],
                "extracted_data": ef.get("extracted_data", {}),
                "filename": ef.get("filename", ""),
            }
            for ef in extracted_files
        ]
        shipment_groups = group_files_into_shipments(group_input)

        await _send_progress(job_id, batch_id, SSEProgress(
            step="grouping",
            shipments_found=len(shipment_groups),
            total=total,
        ))

        # ── Step 4: Build draft shipments + seller intelligence ────
        from backend.services.seller_intelligence import (
            match_or_create_seller,
            apply_seller_defaults,
        )

        await _send_progress(job_id, batch_id, SSEProgress(step="building_drafts", total=len(shipment_groups)))

        for g_idx, group in enumerate(shipment_groups):
            # Gather file data for this group
            group_files = []
            group_file_ids = []
            for ef in extracted_files:
                if str(ef["id"]) in group["file_ids"]:
                    group_files.append({
                        "id": ef["id"],
                        "file_type": ef["file_type"],
                        "extracted_data": ef.get("extracted_data", {}),
                    })
                    group_file_ids.append(ef["id"])

            # Build Xindus-format draft
            shipment_data, confidence_scores = build_draft_shipment(group_files)

            # ── Gaia enrichment: classify items + get tariff data ──
            try:
                from backend.services.gaia_enrichment import enrich_items_with_gaia

                receiver = shipment_data.get("receiver_address") or {}
                dest = normalize_country_code(receiver.get("country") or "", "US")
                shipper = shipment_data.get("shipper_address") or {}
                origin = normalize_country_code(shipper.get("country") or "", "IN")

                await _send_progress(job_id, batch_id, SSEProgress(
                    step="enriching",
                    completed=g_idx,
                    total=len(shipment_groups),
                ))
                shipment_data = await enrich_items_with_gaia(
                    shipment_data, dest, origin,
                )
            except Exception:
                logger.warning(
                    "Gaia enrichment failed for group %d, continuing without",
                    g_idx, exc_info=True,
                )

            # ── Seller intelligence: match/create + apply defaults ──
            seller_id = None
            try:
                shipper_name = (shipment_data.get("shipper_address") or {}).get("name", "")
                shipper_addr = shipment_data.get("shipper_address")
                if shipper_name:
                    seller_id, seller_defaults = await match_or_create_seller(
                        shipper_name, shipper_addr,
                    )
                    if seller_defaults:
                        shipment_data = apply_seller_defaults(
                            shipment_data, seller_defaults, shipper_addr,
                        )
                        logger.info(
                            "Applied seller defaults for '%s' (seller %s)",
                            shipper_name, seller_id,
                        )
            except Exception:
                logger.warning("Seller intelligence failed, continuing without", exc_info=True)

            # Save draft to DB
            draft_id = await db.create_draft(
                batch_id=batch_id,
                shipment_data=shipment_data,
                confidence_scores=confidence_scores,
                grouping_reason=group["reason"],
                seller_id=seller_id,
            )

            # Link files to draft
            await db.link_files_to_draft(draft_id, group_file_ids)

        # ── Done ───────────────────────────────────────────────────
        await db.update_batch_status(batch_id, "review")

        await _send_progress(job_id, batch_id, SSEProgress(
            step="complete",
            batch_id=str(batch_id),
            shipments_found=len(shipment_groups),
            completed=total,
            total=total,
        ))

    except Exception as exc:
        logger.exception("Extraction pipeline failed for batch %s", batch_id)
        await db.update_batch_status(batch_id, "failed")
        await _send_progress(job_id, batch_id, SSEProgress(
            step="error",
            message=str(exc),
        ))
    finally:
        # Signal end of stream
        q = _job_queues.get(job_id)
        if q:
            q.put_nowait(None)


def _estimate_confidence(raw_data: dict[str, Any], file_type: str) -> float:
    """Quick confidence estimate from extracted data."""
    if not raw_data:
        return 0.0

    confidences = []

    def _collect(obj: Any) -> None:
        if isinstance(obj, dict):
            if "confidence" in obj and "value" in obj:
                confidences.append(float(obj["confidence"]))
            else:
                for v in obj.values():
                    _collect(v)
        elif isinstance(obj, list):
            for item in obj:
                _collect(item)

    _collect(raw_data)
    return round(sum(confidences) / len(confidences), 3) if confidences else 0.5


# ---------------------------------------------------------------------------
# GET /api/agent/jobs/{job_id}/stream
# ---------------------------------------------------------------------------


@router.get("/jobs/{job_id}/stream")
async def stream_job_progress(job_id: str):
    """SSE endpoint streaming extraction progress events."""
    q = _job_queues.get(job_id)
    if not q:
        raise HTTPException(404, f"Job {job_id} not found or already completed")

    async def event_generator():
        while True:
            event = await q.get()
            if event is None:
                break
            yield {
                "event": "progress",
                "data": event.model_dump_json(),
            }
        # Cleanup
        _job_queues.pop(job_id, None)

    return EventSourceResponse(event_generator())


# ---------------------------------------------------------------------------
# GET /api/agent/batches/active  (in-flight jobs for persistent status)
# ---------------------------------------------------------------------------


@router.get("/batches/active", response_model=ActiveBatchesResponse)
async def get_active_batches():
    """Return all processing batches with their current step and progress."""
    rows = await db.get_active_batches()
    batches = []
    for r in rows:
        sp = r.get("step_progress") or {}
        if isinstance(sp, str):
            import json as _json
            sp = _json.loads(sp)
        batches.append(ActiveBatch(
            id=r["id"],
            status=r["status"],
            current_step=r.get("current_step"),
            step_progress=sp,
            file_count=r.get("file_count", 0),
            created_at=r.get("created_at"),
        ))
    return ActiveBatchesResponse(batches=batches)


# ---------------------------------------------------------------------------
# GET /api/agent/drafts  (list all drafts with optional status filter)
# ---------------------------------------------------------------------------


@router.get("/drafts", response_model=DraftsListResponse)
async def list_drafts(
    status: str | None = None,
    exclude_status: str | None = None,
    limit: int = 50,
    offset: int = 0,
):
    """List all drafts, optionally filtered by status."""
    if limit < 1:
        limit = 1
    if limit > 200:
        limit = 200
    if offset < 0:
        offset = 0

    drafts_raw, total = await db.get_all_drafts(
        status=status, exclude_status=exclude_status, limit=limit, offset=offset
    )

    # Pre-load seller shipment counts for all drafts with seller_id
    seller_counts: dict[str, int] = {}
    seller_ids_to_load = {d.get("seller_id") for d in drafts_raw if d.get("seller_id")}
    for sid in seller_ids_to_load:
        seller_row = await db.get_seller(sid)
        if seller_row:
            seller_counts[str(sid)] = seller_row.get("shipment_count", 0)

    drafts = []
    for d in drafts_raw:
        # Use corrected_data if available, fall back to shipment_data
        sd = _parse_jsonb(d.get("corrected_data")) or d.get("shipment_data") or {}
        if isinstance(sd, str):
            sd = json.loads(sd)

        draft_file_ids = await db.get_draft_file_ids(d["id"])
        d_seller_id = d.get("seller_id")

        drafts.append(DraftSummary(
            id=d["id"],
            draft_number=d.get("draft_number"),
            status=d["status"],
            file_count=len(draft_file_ids),
            grouping_reason=d.get("grouping_reason"),
            confidence_scores=_parse_jsonb(d.get("confidence_scores")),
            shipper_name=_nested_get(sd, "shipper_address", "name"),
            receiver_name=_nested_get(sd, "receiver_address", "name"),
            box_count=sd.get("total_boxes") or len(sd.get("shipment_boxes", [])),
            total_value=sd.get("total_amount"),
            invoice_number=sd.get("invoice_number"),
            created_at=d.get("created_at"),
            seller_id=d_seller_id,
            seller_shipment_count=seller_counts.get(str(d_seller_id)) if d_seller_id else None,
        ))

    return DraftsListResponse(drafts=drafts, total=total)


# ---------------------------------------------------------------------------
# GET /api/agent/batches/{batch_id}
# ---------------------------------------------------------------------------


@router.get("/batches/{batch_id}", response_model=BatchResponse)
async def get_batch(batch_id: UUID):
    """Get batch metadata with all draft shipments."""
    batch = await db.get_batch(batch_id)
    if not batch:
        raise HTTPException(404, "Batch not found")

    drafts_raw = await db.get_drafts_for_batch(batch_id)
    files_raw = await db.get_files_for_batch(batch_id)

    # Build draft summaries
    drafts = []
    for d in drafts_raw:
        # Use corrected_data if available, fall back to shipment_data
        sd = _parse_jsonb(d.get("corrected_data")) or d.get("shipment_data") or {}
        if isinstance(sd, str):
            sd = json.loads(sd)

        # Count files linked to this draft
        draft_file_ids = await db.get_draft_file_ids(d["id"])

        drafts.append(DraftSummary(
            id=d["id"],
            draft_number=d.get("draft_number"),
            status=d["status"],
            file_count=len(draft_file_ids),
            grouping_reason=d.get("grouping_reason"),
            confidence_scores=_parse_jsonb(d.get("confidence_scores")),
            shipper_name=_nested_get(sd, "shipper_address", "name"),
            receiver_name=_nested_get(sd, "receiver_address", "name"),
            box_count=sd.get("total_boxes") or len(sd.get("shipment_boxes", [])),
            total_value=sd.get("total_amount"),
            invoice_number=sd.get("invoice_number"),
            created_at=d.get("created_at"),
        ))

    files = [
        FileInfo(
            id=f["id"],
            filename=f["filename"],
            file_type=f.get("file_type"),
            page_count=f.get("page_count"),
            confidence=f.get("confidence"),
            processed_at=f.get("processed_at"),
        )
        for f in files_raw
    ]

    return BatchResponse(
        id=batch["id"],
        status=batch["status"],
        file_count=batch["file_count"],
        created_at=batch.get("created_at"),
        completed_at=batch.get("completed_at"),
        drafts=drafts,
        files=files,
    )


# ---------------------------------------------------------------------------
# GET /api/agent/drafts/{draft_id}
# ---------------------------------------------------------------------------


@router.get("/drafts/{draft_id}", response_model=DraftDetail)
async def get_draft(draft_id: UUID):
    """Get full draft details including shipment data and associated files."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")

    files_raw = await db.get_draft_files(draft_id)
    files = [
        FileInfo(
            id=f["id"],
            filename=f["filename"],
            file_type=f.get("file_type"),
            page_count=f.get("page_count"),
            confidence=f.get("confidence"),
            processed_at=f.get("processed_at"),
        )
        for f in files_raw
    ]

    # Load seller profile if available
    seller_id = draft.get("seller_id")
    seller_profile = None
    if seller_id:
        seller_row = await db.get_seller(seller_id)
        if seller_row:
            seller_profile = SellerProfile(
                id=seller_row["id"],
                name=seller_row["name"],
                normalized_name=seller_row["normalized_name"],
                defaults=_parse_jsonb(seller_row.get("defaults")) or {},
                shipper_address=_parse_jsonb(seller_row.get("shipper_address")) or {},
                shipment_count=seller_row.get("shipment_count", 0),
                xindus_customer_id=seller_row.get("xindus_customer_id"),
                created_at=seller_row.get("created_at"),
                updated_at=seller_row.get("updated_at"),
            )

    return DraftDetail(
        id=draft["id"],
        draft_number=draft.get("draft_number"),
        batch_id=draft["batch_id"],
        status=draft["status"],
        shipment_data=_parse_jsonb(draft.get("shipment_data")) or {},
        confidence_scores=_parse_jsonb(draft.get("confidence_scores")),
        grouping_reason=draft.get("grouping_reason"),
        corrected_data=_parse_jsonb(draft.get("corrected_data")),
        xindus_scancode=draft.get("xindus_scancode"),
        files=files,
        created_at=draft.get("created_at"),
        seller_id=seller_id,
        seller=seller_profile,
    )


# ---------------------------------------------------------------------------
# PATCH /api/agent/drafts/{draft_id}
# ---------------------------------------------------------------------------


@router.patch("/drafts/{draft_id}", response_model=DraftDetail)
async def apply_corrections(draft_id: UUID, body: CorrectionRequest):
    """Apply field corrections to a draft shipment."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")

    if draft["status"] not in ("pending_review", "review"):
        raise HTTPException(400, f"Draft is {draft['status']}, cannot apply corrections")

    # Start with existing corrected_data or shipment_data
    current_data = _parse_jsonb(draft.get("corrected_data")) or _parse_jsonb(draft.get("shipment_data")) or {}

    # Get seller_id from draft for scoped corrections
    draft_seller_id = draft.get("seller_id")

    # Apply each correction
    for correction in body.corrections:
        # Special case: xindus_customer_id updates the seller record
        if correction.field_path == "xindus_customer_id":
            if correction.new_value and draft_seller_id:
                await db.update_seller_xindus_customer_id(
                    draft_seller_id, int(correction.new_value)
                )
            elif draft_seller_id:
                await db.update_seller_xindus_customer_id(draft_seller_id, None)
            continue

        # Special case: seller_id updates the DB column, not the shipment JSON
        if correction.field_path == "seller_id":
            new_seller_id = correction.new_value
            if new_seller_id:
                try:
                    seller_uuid = UUID(str(new_seller_id))
                    await db.update_draft_seller(draft_id, seller_uuid)
                    draft_seller_id = seller_uuid
                except (ValueError, TypeError):
                    logger.warning("Invalid seller_id value: %s", new_seller_id)
            else:
                await db.update_draft_seller(draft_id, None)
                draft_seller_id = None
            continue

        _set_nested(current_data, correction.field_path, correction.new_value)

        # Store correction for learning (scoped to seller)
        await db.create_correction(
            draft_id=draft_id,
            field_path=correction.field_path,
            original_value=correction.old_value,
            corrected_value=correction.new_value,
            seller_id=draft_seller_id,
        )

    # Save updated data
    await db.update_draft_corrections(draft_id, current_data)

    return await get_draft(draft_id)


# ---------------------------------------------------------------------------
# POST /api/agent/drafts/{draft_id}/approve
# ---------------------------------------------------------------------------


@router.post("/drafts/{draft_id}/approve", response_model=ApprovalResponse)
async def approve_draft(draft_id: UUID):
    """Approve a draft and push to Xindus B2B API (placeholder)."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")

    if draft["status"] not in ("pending_review",):
        raise HTTPException(400, f"Draft is {draft['status']}, cannot approve")

    # Use corrected_data if available, otherwise shipment_data
    final_data = _parse_jsonb(draft.get("corrected_data")) or _parse_jsonb(draft.get("shipment_data")) or {}

    # TODO: Phase 3 — Push to Xindus API
    # For now, mark as approved with a placeholder scancode
    await db.update_draft_status(draft_id, "approved")

    # Harvest seller defaults from approved shipment
    seller_id = draft.get("seller_id")
    if seller_id:
        try:
            from backend.services.seller_intelligence import harvest_seller_defaults
            await harvest_seller_defaults(seller_id, final_data)
        except Exception:
            logger.warning("Failed to harvest seller defaults", exc_info=True)

    return ApprovalResponse(
        success=True,
        draft_id=draft_id,
        xindus_scancode=None,
        message="Draft approved. Xindus API integration coming in Phase 3.",
    )


# ---------------------------------------------------------------------------
# POST /api/agent/drafts/{draft_id}/submit-xindus
# ---------------------------------------------------------------------------


@router.post("/drafts/{draft_id}/submit-xindus", response_model=SubmissionResult)
async def submit_to_xindus(draft_id: UUID, body: SubmitToXindusRequest):
    """Submit a shipment payload to Xindus UAT and log the result."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")

    from backend.services.xindus_client import submit_b2b_shipment, upload_document
    from backend.storage import s3_key_from_url, download_file as s3_download

    # Upload draft files to Xindus CDN and build "documents" array.
    # Key MUST be "documents" (Xindus @JsonProperty("documents")).
    # Every document MUST have a non-empty CDN URL (validateDocsEmpty rejects empty).
    draft_files = await db.get_draft_files(draft_id)
    documents = []
    packing_uploaded = False
    failed_files = []

    for f in draft_files:
        file_type = f.get("file_type", "invoice") or "invoice"
        if file_type == "invoice":
            doc_type = "invoice"
        elif file_type == "packing_list":
            doc_type = "packinglist"
        elif file_type == "certificate":
            doc_type = "document"
        else:
            doc_type = "document"

        file_url = f.get("file_url", "")
        filename = f.get("filename", "document.pdf")
        file_bytes = None

        # Try S3 (authenticated) first, then HTTP as fallback
        if file_url:
            # Method 1: boto3 S3 client (has auth credentials)
            s3_key = s3_key_from_url(file_url)
            if s3_key:
                try:
                    file_bytes = await s3_download(s3_key)
                    logger.info("S3 download OK for %s (%d bytes)", filename, len(file_bytes))
                except Exception:
                    logger.warning("S3 download failed for %s, trying HTTP", filename)

            # Method 2: HTTP download as fallback (with retries)
            if file_bytes is None:
                for attempt in range(3):
                    try:
                        async with httpx.AsyncClient(timeout=30) as dl_client:
                            dl_resp = await dl_client.get(file_url)
                        if dl_resp.status_code == 200:
                            file_bytes = dl_resp.content
                            logger.info("HTTP download OK for %s (%d bytes)", filename, len(file_bytes))
                            break
                        logger.warning("HTTP download attempt %d failed (%d) for %s",
                                       attempt + 1, dl_resp.status_code, filename)
                    except Exception:
                        logger.warning("HTTP download attempt %d error for %s",
                                       attempt + 1, filename)
                    if attempt < 2:
                        await asyncio.sleep(2 * (attempt + 1))

        if file_bytes is None:
            failed_files.append(filename)
            continue

        # Upload to Xindus CDN
        try:
            cdn_url = await upload_document(file_bytes, filename)
        except Exception:
            logger.warning("Xindus CDN upload failed for %s", filename, exc_info=True)
            failed_files.append(filename)
            continue

        if file_type == "packing_list":
            packing_uploaded = True
        documents.append({
            "id": int(time.time() * 1000) + len(documents),
            "name": doc_type,
            "type": doc_type,
            "url": cdn_url,
            "document_number": "",
        })

    # Fail early if no documents could be uploaded
    if not documents:
        msg = "Could not download files from storage to upload to Xindus."
        if failed_files:
            msg += f" Failed files: {', '.join(failed_files)}"
        msg += " The storage service may be temporarily unavailable — please try again in a few minutes."
        logger.error(msg)
        raise HTTPException(503, msg)

    body.payload["documents"] = documents
    body.payload["invoiceHavePacking"] = not packing_uploaded

    # Create submission record
    submission_id = await db.create_submission(
        draft_id=draft_id,
        environment="uat",
        request_payload=body.payload,
    )

    try:
        http_status, response_body = await submit_b2b_shipment(
            body.payload, body.consignor_id,
        )
    except Exception as exc:
        logger.exception("Xindus submission failed for draft %s", draft_id)
        await db.update_submission_result(
            submission_id,
            http_status=500,
            response_payload={"error": str(exc)},
            status="error",
        )
        return SubmissionResult(
            submission_id=submission_id,
            success=False,
            http_status=500,
            error_description=str(exc),
        )

    # Determine success
    success = 200 <= http_status < 300
    # Xindus wraps data in a list: {"data": [{"scancode": "..."}]}
    raw_data = response_body.get("data")
    data_obj = (raw_data[0] if isinstance(raw_data, list) and raw_data else
                raw_data if isinstance(raw_data, dict) else {})
    scancode = (response_body.get("scancode") or data_obj.get("scancode")) if success else None
    label_b64 = response_body.get("label") or data_obj.get("label")
    error_desc = None
    error_code = None
    if not success:
        error_desc = (
            response_body.get("error_description")
            or response_body.get("message")
            or response_body.get("error")
            or str(response_body)
        )
        error_code = response_body.get("error_code") or response_body.get("code")

    await db.update_submission_result(
        submission_id,
        http_status=http_status,
        response_payload=response_body,
        xindus_scancode=scancode,
        label_base64=label_b64,
        status="success" if success else "error",
    )

    # If successful, update draft status
    if success and scancode:
        await db.update_draft_pushed(draft_id, scancode)

    return SubmissionResult(
        submission_id=submission_id,
        success=success,
        http_status=http_status,
        scancode=scancode,
        error_code=error_code,
        error_description=error_desc,
        response=response_body,
        has_label=bool(label_b64),
    )


# ---------------------------------------------------------------------------
# GET /api/agent/submissions/{submission_id}/label
# ---------------------------------------------------------------------------


@router.get("/submissions/{submission_id}/label")
async def get_submission_label(submission_id: UUID):
    """Return the base64-decoded PDF label for a submission."""
    import base64

    submission = await db.get_submission(submission_id)
    if not submission:
        raise HTTPException(404, "Submission not found")

    label_b64 = submission.get("label_base64")
    if not label_b64:
        raise HTTPException(404, "No label available for this submission")

    pdf_bytes = base64.b64decode(label_b64)
    return FastAPIResponse(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="label_{submission_id}.pdf"'},
    )


# ---------------------------------------------------------------------------
# DELETE /api/agent/drafts/{draft_id}
# ---------------------------------------------------------------------------


@router.delete("/drafts/{draft_id}")
async def reject_draft(draft_id: UUID):
    """Soft-delete (reject) a draft shipment."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")

    await db.update_draft_status(draft_id, "rejected")
    return {"success": True, "message": "Draft rejected"}


# ---------------------------------------------------------------------------
# POST /api/agent/drafts/{draft_id}/archive
# ---------------------------------------------------------------------------


@router.post("/drafts/{draft_id}/archive")
async def archive_draft(draft_id: UUID):
    """Archive a draft shipment."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")
    if draft["status"] == "archived":
        raise HTTPException(400, "Draft is already archived")
    await db.update_draft_status(draft_id, "archived")
    return {"success": True, "message": "Draft archived"}


# ---------------------------------------------------------------------------
# POST /api/agent/drafts/{draft_id}/delete
# ---------------------------------------------------------------------------


@router.post("/drafts/{draft_id}/delete")
async def permanently_delete_draft(draft_id: UUID):
    """Permanently delete a draft shipment from the database."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")
    if draft["status"] not in ("pending_review", "rejected"):
        raise HTTPException(400, f"Cannot delete draft with status '{draft['status']}'")
    await db.delete_draft_permanent(draft_id)
    return {"success": True, "message": "Draft permanently deleted"}


# ---------------------------------------------------------------------------
# POST /api/agent/drafts/{draft_id}/re-extract
# ---------------------------------------------------------------------------


@router.post("/drafts/{draft_id}/re-extract", response_model=DraftDetail)
async def re_extract_draft(draft_id: UUID):
    """Re-run extraction on all files linked to a draft and rebuild shipment data."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")
    if draft["status"] != "pending_review":
        raise HTTPException(400, f"Draft is '{draft['status']}', must be 'pending_review' to re-extract")

    files = await db.get_draft_files(draft_id)
    if not files:
        raise HTTPException(400, "No files linked to this draft")

    from backend.services.llm_extractor import _extract_invoice, _extract_packing_list
    from backend.services.learning import build_extraction_hint
    from backend.services.seller_intelligence import apply_seller_defaults

    # Build correction hints scoped to seller if available
    seller_id = draft.get("seller_id")
    try:
        correction_hints = await build_extraction_hint(seller_id=seller_id)
    except Exception:
        logger.warning("Failed to build extraction hints for re-extract", exc_info=True)
        correction_hints = ""

    extracted_files: list[dict[str, Any]] = []

    for f in files:
        # Download file from storage
        s3_key = f"uploads/{f['batch_id']}/{f['filename']}"
        try:
            file_bytes = await storage.download_file(s3_key)
        except Exception:
            logger.exception("Failed to download file %s for re-extract", f["filename"])
            raise HTTPException(500, f"Failed to download file: {f['filename']}")

        # Process PDF
        pdf_result = await asyncio.to_thread(process_pdf, file_bytes)
        pages = pdf_result.get("pages", [])
        is_vision = any(p.get("image_bytes") for p in pages)

        # Classify
        classification = await classify_document(pages, is_vision=is_vision)
        file_type = classification["document_type"]

        # Extract based on type
        try:
            if file_type == "invoice":
                raw_data = await _extract_invoice(pages, is_vision, correction_hints=correction_hints)
            elif file_type == "packing_list":
                raw_data = await _extract_packing_list(pages, is_vision, correction_hints=correction_hints)
            else:
                raw_data = await _extract_invoice(pages, is_vision, correction_hints=correction_hints)
        except Exception:
            logger.exception("Re-extraction failed for %s", f["filename"])
            raw_data = {}

        confidence = _estimate_confidence(raw_data, file_type)

        # Update file record
        await db.update_file_extraction(
            f["id"],
            file_type=file_type,
            page_count=len(pages),
            extracted_data=raw_data,
            confidence=confidence,
        )

        extracted_files.append({
            "id": f["id"],
            "file_type": file_type,
            "extracted_data": raw_data,
        })

    # Build merged shipment data
    shipment_data, confidence_scores = build_draft_shipment(extracted_files)

    # Re-apply seller defaults
    if seller_id:
        try:
            seller = await db.get_seller(seller_id)
            if seller:
                seller_defaults = _parse_jsonb(seller.get("defaults"))
                shipper_addr = _parse_jsonb(seller.get("shipper_address"))
                if seller_defaults:
                    shipment_data = apply_seller_defaults(shipment_data, seller_defaults, shipper_addr)
        except Exception:
            logger.warning("Failed to apply seller defaults during re-extract", exc_info=True)

    # Save rebuilt data (clears corrected_data)
    await db.update_draft_shipment_data(draft_id, shipment_data, confidence_scores)

    return await get_draft(draft_id)


# ---------------------------------------------------------------------------
# POST /api/agent/drafts/{draft_id}/classify  (Gaia classification)
# ---------------------------------------------------------------------------


@router.post("/drafts/{draft_id}/classify", response_model=DraftDetail)
async def classify_draft(draft_id: UUID):
    """Run Gaia classification on all items in a draft."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")

    from backend.services.gaia_enrichment import classify_draft_items

    result = await classify_draft_items(draft_id)
    if result is None:
        raise HTTPException(500, "Classification failed")

    return await get_draft(draft_id)


# ---------------------------------------------------------------------------
# POST /api/agent/tariff-lookup  (single HTS code tariff detail)
# ---------------------------------------------------------------------------


@router.post("/tariff-lookup")
async def tariff_lookup(body: dict):
    """Look up tariff detail for a single HTS code.

    Body: {tariff_code: str, destination_country: str, origin_country: str}
    Returns duty breakdown with base rate, cumulative rate, and scenarios.
    """
    tariff_code = (body.get("tariff_code") or "").strip()
    destination = normalize_country_code(body.get("destination_country") or "", "US")
    origin = normalize_country_code(body.get("origin_country") or "", "IN")

    if not tariff_code:
        raise HTTPException(400, "tariff_code is required")

    from backend.services.gaia_client import get_tariff_detail, calculate_cumulative_duty

    tariff_data = await get_tariff_detail(destination, tariff_code, origin)
    if not tariff_data:
        raise HTTPException(404, f"No tariff data for {tariff_code} ({origin}→{destination})")

    base_rate, cumulative_rate, scenarios = calculate_cumulative_duty(tariff_data)

    # Extract remedy flags
    remedy_flags: dict = {}
    for flag in tariff_data.get("flags") or []:
        if flag.get("name") == "remedy":
            val = flag.get("value") or {}
            remedy_flags = {
                "add_risk": bool(val.get("possible_add_required_indicator")),
                "cvd_risk": bool(val.get("possible_cvd_duty_required_indicator")),
            }

    return {
        "duty_rate": cumulative_rate,
        "base_duty_rate": base_rate,
        "tariff_scenarios": scenarios,
        "remedy_flags": remedy_flags,
    }


# ---------------------------------------------------------------------------
# POST /api/agent/drafts/{draft_id}/files  (add files to draft)
# ---------------------------------------------------------------------------


@router.post("/drafts/{draft_id}/files", response_model=DraftDetail)
async def add_files_to_draft(draft_id: UUID, files: list[UploadFile] = File(...)):
    """Upload new files to an existing draft, then re-extract."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")
    if draft["status"] != "pending_review":
        raise HTTPException(400, f"Draft is '{draft['status']}', must be 'pending_review' to add files")
    if not files:
        raise HTTPException(400, "No files uploaded")

    batch_id = draft["batch_id"]

    # Validate and upload each file
    new_file_ids: list[UUID] = []
    for f in files:
        ext = "." + f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(400, f"Unsupported file type: {f.filename}")

        content = await f.read()
        if len(content) > MAX_FILE_SIZE_MB * 1024 * 1024:
            raise HTTPException(400, f"File too large: {f.filename} (max {MAX_FILE_SIZE_MB}MB)")

        s3_key = f"uploads/{batch_id}/{f.filename}"
        file_url = await storage.upload_file(s3_key, content)
        file_id = await db.create_file(batch_id, f.filename, file_url)
        new_file_ids.append(file_id)

    # Link new files to draft
    await db.link_files_to_draft(draft_id, new_file_ids)

    # Increment batch file_count
    pool = db.get_pool()
    await pool.execute(
        "UPDATE upload_batches SET file_count = file_count + $2 WHERE id = $1",
        batch_id, len(new_file_ids),
    )

    # Re-extract with all files (existing + new)
    return await re_extract_draft(draft_id)


# ---------------------------------------------------------------------------
# DELETE /api/agent/drafts/{draft_id}/files/{file_id}  (remove file from draft)
# ---------------------------------------------------------------------------


@router.delete("/drafts/{draft_id}/files/{file_id}", response_model=DraftDetail)
async def remove_file_from_draft(draft_id: UUID, file_id: UUID):
    """Remove a file from a draft (unlink only), then re-extract."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")
    if draft["status"] != "pending_review":
        raise HTTPException(400, f"Draft is '{draft['status']}', must be 'pending_review' to remove files")

    file_record = await db.get_file(file_id)
    if not file_record:
        raise HTTPException(404, "File not found")

    linked_ids = await db.get_draft_file_ids(draft_id)
    if file_id not in linked_ids:
        raise HTTPException(400, "File is not linked to this draft")
    if len(linked_ids) <= 1:
        raise HTTPException(400, "Cannot remove the last file from a draft")

    await db.unlink_file_from_draft(draft_id, file_id)

    # Re-extract with remaining files
    return await re_extract_draft(draft_id)


# ---------------------------------------------------------------------------
# GET /api/agent/drafts/{draft_id}/files/{file_id}/download
# ---------------------------------------------------------------------------


@router.get("/drafts/{draft_id}/files/{file_id}/download")
async def download_draft_file(draft_id: UUID, file_id: UUID):
    """Download the original PDF for a file linked to a draft."""
    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")

    file_record = await db.get_file(file_id)
    if not file_record:
        raise HTTPException(404, "File not found")

    linked_ids = await db.get_draft_file_ids(draft_id)
    if file_id not in linked_ids:
        raise HTTPException(400, "File is not linked to this draft")

    s3_key = f"uploads/{file_record['batch_id']}/{file_record['filename']}"
    try:
        file_bytes = await storage.download_file(s3_key)
    except Exception:
        logger.exception("Failed to download file %s", file_record["filename"])
        raise HTTPException(500, f"Failed to download file: {file_record['filename']}")

    filename = file_record["filename"]
    return FastAPIResponse(
        content=file_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# GET /api/agent/drafts/{draft_id}/download
# ---------------------------------------------------------------------------


@router.get("/drafts/{draft_id}/download")
async def download_draft_excel(draft_id: UUID, format: str = "summary"):
    """Download the draft shipment data as an Excel workbook.

    Query params:
        format: "summary" (default, 3-sheet overview) or "xpressb2b" (21-col upload sheet)
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    draft = await db.get_draft(draft_id)
    if not draft:
        raise HTTPException(404, "Draft not found")

    data = _parse_jsonb(draft.get("corrected_data")) or _parse_jsonb(draft.get("shipment_data"))
    if not data:
        raise HTTPException(400, "No shipment data in draft")

    inv = data.get("invoice_number", str(draft_id)[:8])
    safe_inv = "".join(c if c.isalnum() or c in "-_." else "_" for c in str(inv))

    # --- Shared styling ---
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    def _auto_fit(ws, min_w=10, max_w=40):
        for ci in range(1, ws.max_column + 1):
            mx = 0
            for row in ws.iter_rows(min_col=ci, max_col=ci, values_only=False):
                for cell in row:
                    if cell.value is not None:
                        mx = max(mx, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max(mx + 2, min_w), max_w)

    boxes = data.get("shipment_boxes", []) or []

    # =====================================================================
    # FORMAT: XpressB2B Multi Address (exact 21-column Xindus upload sheet)
    # =====================================================================
    if format == "xpressb2b":
        COLUMNS = [
            "Box Number",
            "Receiver Name",
            "Receiver Address",
            "Receiver City",
            "Receiver Zip",
            "Receiver State",
            "Receiver Country",
            "Receiver Phone Number",
            "Receiver Extension No",
            "Receiver Email",
            "Box Length (cms)",
            "Box Width (cms)",
            "Box Height (cms)",
            "Box Weight (kgs)",
            "Item Description",
            "Item Qty",
            "Item Unit Weight Kg",
            "HS Code (Origin)",
            "HS Code (Destination)",
            "Item Unit Price",
            "IGST % (For GST taxtype)",
        ]
        BOX_COL_COUNT = 14  # cols 1-14 are box-level (only on first item row)

        wb = Workbook()
        ws = wb.active
        ws.title = "XpressB2B Multi Address"

        # Header row
        for ci, h in enumerate(COLUMNS, start=1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = hdr_align
            c.border = thin_border

        # Fallback receiver from top-level
        top_recv = data.get("receiver_address", {}) or {}

        current_row = 2
        for box in boxes:
            items = box.get("shipment_box_items", []) or [{}]
            recv = box.get("receiver_address", {}) or {}
            # Use box receiver if it has a name, else fallback to top-level
            if not recv.get("name"):
                recv = top_recv

            for j, item in enumerate(items):
                is_first = j == 0

                if is_first:
                    row_data = [
                        box.get("box_id", ""),
                        recv.get("name", ""),
                        recv.get("address", ""),
                        recv.get("city", ""),
                        recv.get("zip", ""),
                        recv.get("state", ""),
                        recv.get("country", ""),
                        recv.get("phone", ""),
                        recv.get("extension_number", ""),
                        recv.get("email", ""),
                        box.get("length", ""),
                        box.get("width", ""),
                        box.get("height", ""),
                        box.get("weight", ""),
                    ]
                else:
                    row_data = [""] * BOX_COL_COUNT

                # Item-level columns (15-21)
                row_data.extend([
                    item.get("description", ""),
                    item.get("quantity", ""),
                    item.get("weight", ""),
                    item.get("ehsn", ""),
                    item.get("ihsn", ""),
                    item.get("unit_price", ""),
                    item.get("igst_amount", ""),
                ])

                for ci, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=ci, value=val if val else "")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

                current_row += 1

        _auto_fit(ws)
        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"XpressB2B_{safe_inv}.xlsx"

        return FastAPIResponse(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # =====================================================================
    # FORMAT: Summary (default 3-sheet overview)
    # =====================================================================
    wb = Workbook()

    # ---- Sheet 1: Shipment Info ----
    ws = wb.active
    ws.title = "Shipment"
    info_fields = [
        ("Invoice Number", data.get("invoice_number", "")),
        ("Invoice Date", data.get("invoice_date", "")),
        ("Shipping Method", data.get("shipping_method", "")),
        ("Terms of Trade", data.get("terms_of_trade", "")),
        ("Tax Type", data.get("tax_type", "")),
        ("Origin Clearance", data.get("origin_clearance_type", "")),
        ("Destination Clearance", data.get("destination_clearance_type", "")),
        ("Country", data.get("country", "")),
        ("Port of Entry", data.get("port_of_entry", "")),
        ("Destination CHA", data.get("destination_cha", "")),
        ("Purpose", data.get("purpose_of_booking", "")),
        ("Export Reference", data.get("export_reference", "")),
    ]
    for label, addr_key in [
        ("Shipper", "shipper_address"),
        ("Receiver", "receiver_address"),
        ("Billing", "billing_address"),
        ("IOR", "ior_address"),
    ]:
        addr = data.get(addr_key, {}) or {}
        parts = [addr.get("name", ""), addr.get("address", ""),
                 addr.get("city", ""), addr.get("state", ""),
                 addr.get("zip", ""), addr.get("country", "")]
        info_fields.append((f"{label} Address", ", ".join(p for p in parts if p)))
        if addr.get("phone"):
            info_fields.append((f"{label} Phone", addr["phone"]))
        if addr.get("email"):
            info_fields.append((f"{label} Email", addr["email"]))

    for i, (label, value) in enumerate(info_fields, start=1):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(bold=True, size=10)
        lc.border = thin_border
        vc = ws.cell(row=i, column=2, value=str(value) if value else "")
        vc.font = data_font
        vc.border = thin_border
        vc.alignment = wrap
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    # ---- Sheet 2: Boxes & Items ----
    ws2 = wb.create_sheet("Boxes & Items")
    box_headers = [
        "Box #", "Length", "Width", "Height", "Weight (kg)", "UOM",
        "Battery", "Receiver Name", "Receiver City", "Receiver Country",
        "Item Description", "Qty", "Unit Price", "Total Price",
        "Export HSN", "Import HSN", "Origin", "FOB Value", "IGST %", "Category",
    ]
    for ci, h in enumerate(box_headers, start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    row = 2
    for box in boxes:
        items = box.get("shipment_box_items", []) or [{}]
        recv = box.get("receiver_address", {}) or {}
        for j, item in enumerate(items):
            is_first = j == 0
            ws2.cell(row=row, column=1, value=box.get("box_id", "") if is_first else "").border = thin_border
            ws2.cell(row=row, column=2, value=box.get("length", 0) if is_first else "").border = thin_border
            ws2.cell(row=row, column=3, value=box.get("width", 0) if is_first else "").border = thin_border
            ws2.cell(row=row, column=4, value=box.get("height", 0) if is_first else "").border = thin_border
            ws2.cell(row=row, column=5, value=box.get("weight", 0) if is_first else "").border = thin_border
            ws2.cell(row=row, column=6, value=box.get("uom", "cm") if is_first else "").border = thin_border
            ws2.cell(row=row, column=7, value="Yes" if box.get("has_battery") else "No" if is_first else "").border = thin_border
            ws2.cell(row=row, column=8, value=recv.get("name", "") if is_first else "").border = thin_border
            ws2.cell(row=row, column=9, value=recv.get("city", "") if is_first else "").border = thin_border
            ws2.cell(row=row, column=10, value=recv.get("country", "") if is_first else "").border = thin_border
            ws2.cell(row=row, column=11, value=item.get("description", "")).border = thin_border
            ws2.cell(row=row, column=12, value=item.get("quantity", "")).border = thin_border
            ws2.cell(row=row, column=13, value=item.get("unit_price", "")).border = thin_border
            ws2.cell(row=row, column=14, value=item.get("total_price", "")).border = thin_border
            ws2.cell(row=row, column=15, value=item.get("ehsn", "")).border = thin_border
            ws2.cell(row=row, column=16, value=item.get("ihsn", "")).border = thin_border
            ws2.cell(row=row, column=17, value=item.get("country_of_origin", "")).border = thin_border
            ws2.cell(row=row, column=18, value=item.get("unit_fob_value", "")).border = thin_border
            ws2.cell(row=row, column=19, value=item.get("igst_amount", "")).border = thin_border
            ws2.cell(row=row, column=20, value=item.get("category", "")).border = thin_border
            row += 1

    for ci in range(1, 21):
        ws2.column_dimensions[chr(64 + ci) if ci <= 26 else ""].width = 14
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["K"].width = 30

    # ---- Sheet 3: Product Details (customs) ----
    ws3 = wb.create_sheet("Product Details")
    pd_headers = ["Description", "HSN Code", "Value"]
    for ci, h in enumerate(pd_headers, start=1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = thin_border
    for pi, pd in enumerate(data.get("product_details", []) or [], start=2):
        ws3.cell(row=pi, column=1, value=pd.get("product_description", "")).border = thin_border
        ws3.cell(row=pi, column=2, value=pd.get("hsn_code", "")).border = thin_border
        ws3.cell(row=pi, column=3, value=pd.get("value", "")).border = thin_border
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 14

    # ---- Save to bytes ----
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"draft_{safe_inv}.xlsx"

    return FastAPIResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# GET /api/agent/shipping-methods
# ---------------------------------------------------------------------------


@router.get("/shipping-methods")
async def list_shipping_methods(b2b: bool = True):
    methods = await db.get_shipping_methods(b2b_only=b2b)
    return methods


# ---------------------------------------------------------------------------
# GET /api/agent/sellers  (list all sellers)
# GET /api/agent/sellers/match?name=...  (match a seller by name)
# ---------------------------------------------------------------------------


@router.get("/sellers", response_model=SellersListResponse)
async def list_sellers():
    """List all known seller profiles."""
    rows = await db.get_all_sellers()
    sellers = [
        SellerProfile(
            id=r["id"],
            name=r["name"],
            normalized_name=r["normalized_name"],
            defaults=_parse_jsonb(r.get("defaults")) or {},
            shipper_address=_parse_jsonb(r.get("shipper_address")) or {},
            shipment_count=r.get("shipment_count", 0),
            xindus_customer_id=r.get("xindus_customer_id"),
            created_at=r.get("created_at"),
            updated_at=r.get("updated_at"),
        )
        for r in rows
    ]
    return SellersListResponse(sellers=sellers)


@router.get("/sellers/{seller_id}/history")
async def get_seller_history(seller_id: UUID):
    """Return deduplicated products, billing, IOR, and receiver addresses from approved shipments."""
    seller = await db.get_seller(seller_id)
    if not seller:
        raise HTTPException(404, "Seller not found")

    drafts = await db.get_approved_drafts_for_seller(seller_id)

    seen_products: set[str] = set()
    seen_billing: set[str] = set()
    seen_ior: set[str] = set()
    seen_receiver: set[str] = set()

    products: list[dict] = []
    billing_addresses: list[dict] = []
    ior_addresses: list[dict] = []
    receiver_addresses: list[dict] = []

    for d in drafts:
        sd = _parse_jsonb(d.get("corrected_data")) or _parse_jsonb(d.get("shipment_data")) or {}

        # Products
        for p in sd.get("product_details", []) or []:
            desc = (p.get("product_description") or "").strip()
            hsn = (p.get("hsn_code") or "").strip()
            if not desc:
                continue
            key = f"{desc.lower()}|{hsn.lower()}"
            if key not in seen_products:
                seen_products.add(key)
                products.append({"product_description": desc, "hsn_code": hsn, "value": p.get("value", 0)})

        # Billing address
        ba = sd.get("billing_address") or {}
        if ba.get("name"):
            bkey = f"{ba['name'].lower()}|{(ba.get('city') or '').lower()}|{(ba.get('country') or '').lower()}"
            if bkey not in seen_billing:
                seen_billing.add(bkey)
                billing_addresses.append(ba)

        # IOR address
        ia = sd.get("ior_address") or {}
        if ia.get("name"):
            ikey = f"{ia['name'].lower()}|{(ia.get('city') or '').lower()}|{(ia.get('country') or '').lower()}"
            if ikey not in seen_ior:
                seen_ior.add(ikey)
                ior_addresses.append(ia)

        # Receiver addresses (from boxes + top-level)
        for addr_src in [sd.get("receiver_address")] + [
            b.get("receiver_address") for b in (sd.get("shipment_boxes") or [])
        ]:
            if not addr_src or not addr_src.get("name"):
                continue
            rkey = f"{addr_src['name'].lower()}|{(addr_src.get('city') or '').lower()}|{(addr_src.get('country') or '').lower()}"
            if rkey not in seen_receiver:
                seen_receiver.add(rkey)
                receiver_addresses.append(addr_src)

    return {
        "products": products,
        "billing_addresses": billing_addresses,
        "ior_addresses": ior_addresses,
        "receiver_addresses": receiver_addresses,
    }


@router.get("/sellers/match", response_model=SellerProfile)
async def match_seller(name: str):
    """Match a seller by name (exact + fuzzy)."""
    from backend.services.seller_intelligence import match_or_create_seller

    if not name.strip():
        raise HTTPException(400, "name parameter required")

    seller_id, _ = await match_or_create_seller(name)
    seller_row = await db.get_seller(seller_id)
    if not seller_row:
        raise HTTPException(404, "Seller not found")

    return SellerProfile(
        id=seller_row["id"],
        name=seller_row["name"],
        normalized_name=seller_row["normalized_name"],
        defaults=_parse_jsonb(seller_row.get("defaults")) or {},
        shipper_address=_parse_jsonb(seller_row.get("shipper_address")) or {},
        shipment_count=seller_row.get("shipment_count", 0),
        xindus_customer_id=seller_row.get("xindus_customer_id"),
        created_at=seller_row.get("created_at"),
        updated_at=seller_row.get("updated_at"),
    )


# ---------------------------------------------------------------------------
# Xindus customer/address endpoints (mirrored data)
# ---------------------------------------------------------------------------


@router.get("/xindus/customers/search")
async def search_xindus_customers(q: str = "", limit: int = 20):
    """Search Xindus customers by company name or ID."""
    if not q.strip() or len(q.strip()) < 2:
        return {"customers": []}
    rows = await db.search_xindus_customers(q.strip(), min(limit, 50))
    customers = [XindusCustomer(**r) for r in rows]
    return {"customers": [c.model_dump() for c in customers]}


@router.get("/xindus/customers/{customer_id}")
async def get_xindus_customer(customer_id: int):
    """Get a single Xindus customer by ID."""
    row = await db.get_xindus_customer(customer_id)
    if not row:
        raise HTTPException(404, "Customer not found")
    return XindusCustomer(**row).model_dump()


@router.get("/xindus/customers/{customer_id}/addresses")
async def get_xindus_addresses(customer_id: int, type: str | None = None):
    """Get addresses for a Xindus customer, optionally filtered by type."""
    rows = await db.get_xindus_addresses(customer_id, type)
    addresses = [XindusAddress(**r) for r in rows]
    return {"addresses": [a.model_dump() for a in addresses]}


@router.post("/xindus/sync", response_model=XindusSyncResponse)
async def sync_xindus_data(body: XindusSyncRequest):
    """Bulk upsert Xindus customers and addresses (called by local sync script)."""
    customers_count = await db.upsert_xindus_customers(body.customers)
    addresses_count = await db.upsert_xindus_addresses(body.addresses)
    return XindusSyncResponse(
        customers_upserted=customers_count,
        addresses_upserted=addresses_count,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_jsonb(val: Any) -> dict[str, Any] | None:
    """Parse a JSONB value that may be a string or already a dict."""
    if val is None:
        return None
    if isinstance(val, dict):
        return val
    if isinstance(val, str):
        try:
            return json.loads(val)
        except (json.JSONDecodeError, TypeError):
            return None
    return None


def _nested_get(data: dict[str, Any], *keys: str) -> str | None:
    """Safely get a nested value from a dict."""
    obj = data
    for k in keys:
        if not isinstance(obj, dict):
            return None
        obj = obj.get(k)
    if obj is None:
        return None
    return str(obj) if obj else None


def _set_nested(data: dict[str, Any], path: str, value: Any) -> None:
    """Set a value in a nested dict using dot-separated path.

    Supports array indices: "shipment_boxes.0.gross_weight_kg"
    """
    keys = path.split(".")
    obj = data
    for key in keys[:-1]:
        if key.isdigit():
            idx = int(key)
            if isinstance(obj, list) and 0 <= idx < len(obj):
                obj = obj[idx]
            else:
                return
        else:
            if not isinstance(obj, dict):
                return
            if key not in obj:
                obj[key] = {}
            obj = obj[key]

    final_key = keys[-1]
    if final_key.isdigit() and isinstance(obj, list):
        idx = int(final_key)
        if 0 <= idx < len(obj):
            obj[idx] = value
    elif isinstance(obj, dict):
        obj[final_key] = value
